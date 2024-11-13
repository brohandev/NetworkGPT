import os
from prettyprinter import pprint as pp
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.realpath(__file__))

# Controller API Base Paths
DNAC_KB_INPUT = BASE_DIR + "/documents/DNAC_API_Knowledge_Base.xlsx"
VMANAGE_KB_INPUT = BASE_DIR + "/documents/vManage_API_Knowledge_Base.xlsx"
ISE_KB_INPUT = BASE_DIR + "/documents/ISE_API_Knowledge_Base.xlsx"
OPENAI_KB_INPUT = BASE_DIR + "/documents/OpenAI_Prompt_Knowledge_Base.xlsx"

# Knowledge Base Paths
DNAC_KB_PATH = BASE_DIR + "/documents/DNAC_KB.json"
VMANAGE_KB_PATH = BASE_DIR + "/documents/VMANAGE_KB.json"
ISE_KB_PATH = BASE_DIR + "/documents/ISE_KB.json"
STEALTHWATCH_KB_PATH = BASE_DIR + "/documents/STEALTHWATCH_KB.json"
DEVICE_DOMAIN_MAPPING_PATH = BASE_DIR + "/documents/DEVICE_DOMAIN_MAPPING.json"


def openAI_prompt_knowledge_base_to_dict(doc, sheet_name):
    workbook = load_workbook(doc)
    sheet = workbook[sheet_name]
    prompt_list = []

    try:
        for row in range(2, sheet.max_row + 1):
            prompt_identifier = str(sheet[row][0].value)
            prompt_description = str(sheet[row][1].value)
            prompt_list.append({
                "role": prompt_identifier,
                "content": prompt_description.strip()
            })

    except Exception as e:
        print(str(e))

    return prompt_list


def category_kb_to_dict_list(doc):
    workbook = load_workbook(doc)
    category_kb = []
    category_list = []

    try:
        sheet = workbook["INTENT_CATEGORIES"]
        for row in range(2, sheet.max_row + 1):
            category_name = str(sheet[row][0].value)
            category_desc = str(sheet[row][1].value)

            category_kb.append({
                "category": category_name,
                "description": category_desc
            })
            category_list.append(category_name)

    except Exception as e:
        print(str(e))

    return category_kb, category_list


def api_knowledge_base_to_dict(document, sheet_name):
    workbook = load_workbook(document)
    sheet_name_list = ["INTENT_CATEGORIES"]
    api_list = []

    try:
        sheet = workbook[sheet_name]
        for row in range(2, sheet.max_row + 1):
            category_name = str(sheet[row][0].value)
            category_desc = str(sheet[row][1].value)
            api_list.append({
                "category": category_name,
                "description": category_desc
            })

    except Exception as e:
        print(str(e))

    return api_list


if __name__ == '__main__':
    # api_list = api_knowledge_base_to_dict(DNAC_KB_INPUT)
    # prompt_list = openAI_prompt_knowledge_base_to_dict(OPENAI_KB_INPUT, "ROLES")
    # pp(prompt_list)
    print(DNAC_KB_PATH)
