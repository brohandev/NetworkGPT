# package import
import json
import logging
from openpyxl import load_workbook
from requests import (
    post, Session,
    ConnectionError, HTTPError, Timeout
)
from webex_bot.models.command import Command
from webex_bot.webex_bot import WebexBot

# local file import
from Auxiliary.helper import (
    substring_exists,
    write_to_json
)
from Authentication.credentials import (
    OPENAI_API_KEY,
    OPENAI_COMPLETION_URL,
    OPENAI_MODEL,
    OPENAI_TEMPERATURE,
    WEBEX_BOT_ACCESS_TOKEN,
    WEBEX_BOT_NAME
)
from Storage.filepaths import (
    dnac_kb_filepath,
    ise_kb_filepath,
    prompt_kb_filepath,
    snam_kb_filepath,
    vmanage_kb_filepath,
    knox_kb_filepath,
    device_domain_map_filepath
)

logging.basicConfig()
log = logging.getLogger("Webex_Chatbot_Operation")
log.setLevel(logging.INFO)


class AnswerCommand(Command):
    def __init__(self):
        super().__init__()

    def execute(self, message, attachment_actions=None, activity=None):
        """
        Process incoming message from the user, and return the text answer back to the user.

        :param message: message with command already stripped from the user
        :param attachment_actions: attachment_actions object (N.A)
        :param activity: activity object (N.A)

        :return: a string or Response object (or a list of either). Use Response if you want to return another card.
        """

        return Chatbot().handle_message(message)


class Chatbot:
    def __init__(self):
        # instantiate OpenAI session
        self.open_ai_session = self.open_ai_authenticate()
        # instantiate Webex bot
        self.chatbot = WebexBot(teams_bot_token=WEBEX_BOT_ACCESS_TOKEN,
                                bot_name=WEBEX_BOT_NAME,
                                # approved_rooms=[
                                #     WEBEX_RODEV_ROOM
                                # ],
                                include_demo_commands=False)
        # Clear Webex bot default commands
        self.chatbot.commands.clear()
        # Add custom command, and set it as the new default command
        self.chatbot.help_command = AnswerCommand()
        # initialize bot and chat history with system prompt
        self.system = json.dumps(self.prompt_kb_to_dict(doc=prompt_kb_filepath, sheet_name="SYSTEM_INIT")[0])
        self.chat_history = [{"role": "system", "content": self.system}]

    @staticmethod
    def open_ai_authenticate():
        """
        Creates a Session object and authenticates with the OpenAI servers. A POST call is made to test if the request
        complies with OpenAI's accepted API syntax. Upon successful status code, the Session object is updated with the
        Bearer Token in the header and returned to the caller. Upon failure during the previous step, the program stops
        and exits.
        """
        response = None
        session = Session()

        try:
            headers = {
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json",
                "Accept": "application/json"
            }
            response = post(url=OPENAI_COMPLETION_URL,
                            headers=headers,
                            data=json.dumps({
                                "model": OPENAI_MODEL,
                                "messages": [{
                                    "role": "user",
                                    "content": "Say this is a test!"
                                }],
                                "temperature": OPENAI_TEMPERATURE
                            }),
                            verify=False)

            if response.ok and response:
                session.headers.update(headers)
                log.info("OpenAI Authentication: Successful.")
            else:
                response.raise_for_status()

        except HTTPError:
            if response.status_code == 403:
                log.error("OpenAI Authentication: HTTP 401: Invalid API key. Refresh API key.")
                exit()
        except ConnectionError:
            log.error("OpenAI Authentication: Connection: Check network connectivity to outbound Internet or check "
                      "URL validity.")
            exit()
        except Timeout:
            log.error("OpenAI Authentication: Timeout: Re-attempt authentication method.")
            exit()
        except Exception as e:
            log.error(f"OpenAI Authentication: Unknown exception: Deeper troubleshooting required to fix {e}")
            exit()

        return session

    def prompt_kb_to_dict(self, doc, sheet_name):
        workbook = load_workbook(doc)
        sheet = workbook[sheet_name]
        prompt_list = []

        try:
            for row in range(2, sheet.max_row + 1):
                prompt_identifier = str(sheet[row][0].value)
                prompt_description = str(sheet[row][1].value)
                prompt_list.append({
                    "role": prompt_identifier,
                    "content": prompt_description.strip()
                })

        except Exception as e:
            log.error(f"Prompt KB Ingestion: Unknown exception: Deeper troubleshooting required to fix {e}")
            exit()

        return prompt_list

    def category_kb_to_dict(self, doc):
        workbook = load_workbook(filename=doc)
        category_kb = []
        category_list = []

        try:
            sheet = workbook["INTENT_CATEGORIES"]
            for row in range(2, sheet.max_row + 1):
                category_name = str(sheet[row][0].value)
                category_desc = str(sheet[row][1].value)

                category_kb.append({
                    "category": category_name,
                    "description": category_desc
                })
                category_list.append(category_name)

        except Exception as e:
            print(str(e))

        return category_kb, category_list

    def knowledge_base_segmentor(self, intent):
        kb = None
        with open(dnac_kb_filepath, "r") as file:
            dnac_json_data = json.load(file)
        with open(vmanage_kb_filepath, "r") as file:
            vmanage_json_data = json.load(file)
        with open(ise_kb_filepath, "r") as file:
            ise_json_data = json.load(file)
        with open(knox_kb_filepath, "r") as file:
            knox_json_data = json.load(file)
        with open(snam_kb_filepath, "r") as file:
            snam_json_data = json.load(file)

        match intent:
            case "LAN_DEVICES":
                kb = dnac_json_data["LAN_DEVICES"]
            case "LAN_INTERFACES":
                kb = dnac_json_data["LAN_INTERFACES"]
            case "LAN_CLIENTS":
                kb = dnac_json_data["LAN_CLIENTS"]
            case "LAN_ISSUES":
                kb = dnac_json_data["LAN_ISSUES"]
                if len(kb) > 10:
                    kb = kb[:10]
            case "WAN_DEVICES":
                kb = vmanage_json_data["WAN_DEVICES"]
            case "WAN_INTERFACES":
                kb = vmanage_json_data["WAN_INTERFACES"]
            case "WAN_ISSUES":
                kb = vmanage_json_data["WAN_ISSUES"]
                if len(kb) > 10:
                    kb = kb[:10]
            case "ISE_AUTHENTICATION":
                kb = ise_json_data["AUTHENTICATION_POLICIES"]
            case "ISE_AUTHORIZATION":
                kb = ise_json_data["AUTHORIZATION_POLICIES"]
            case "STEALTHWATCH_ISSUES":
                kb = snam_json_data["STEALTHWATCH_ALARMS"]
                if len(kb) > 10:
                    kb = kb[:10]
            case "KNOX_DEVICES":
                kb = knox_json_data["SAMSUNG_DEVICES"]
            case "OVERALL_ISSUES":
                dnac_issues = []
                vmanage_issues = []
                for issue in dnac_json_data["LAN_ISSUES"]:
                    if type(issue) == dict:
                        issue['domain'] = 'LAN'
                    dnac_issues.append(issue)
                for issue in vmanage_json_data["WAN_ISSUES"]:
                    if type(issue) == dict:
                        issue['domain'] = 'WAN'
                    vmanage_issues.append(issue)
                if len(dnac_issues) > 10:
                    dnac_issues = dnac_issues[:10]
                if len(vmanage_issues) > 10:
                    vmanage_issues = vmanage_issues[:10]
                kb = dnac_issues + vmanage_issues
        return json.dumps(kb)

    def generate_device_domain_mapping(self):
        try:
            with open(dnac_kb_filepath, "r") as file:
                dnac_json_data = json.load(file)
            with open(vmanage_kb_filepath, "r") as file:
                vmanage_json_data = json.load(file)
            with open(knox_kb_filepath, "r") as file:
                knox_json_data = json.load(file)

            device_domain_mapping = {}
            lan_devices = dnac_json_data["LAN_DEVICES"]
            for device in lan_devices:
                device_domain_mapping[device["hostname"]] = "LAN"
            wan_devices = vmanage_json_data["WAN_DEVICES"]
            for device in wan_devices:
                device_domain_mapping[device["host-name"]] = "WAN"
            knox_devices = knox_json_data["SAMSUNG_DEVICES"]
            for device in knox_devices:
                if len(device["model"]):
                    device_domain_mapping[device["model"]] = "SAMSUNG"

            write_to_json(
                document=device_domain_map_filepath,
                content=device_domain_mapping
            )
            log.info("Device Domain Mapping: Update Successful")
        except:
            log.info("Device Domain Mapping: Update Unsuccessful")

    def discover_intent(self, question):
        # Detect device in question, and append domain to the end of question
        with open(device_domain_map_filepath, "r") as file:
            device_domain_mapping = json.load(file)
        for hostname, domain in device_domain_mapping.items():
            if substring_exists(hostname.lower(), question.lower()):
                question += f" '{hostname}' is in {domain}."

        # Ask ChatGPT which category the user's question falls under
        user_prompt = self.prompt_kb_to_dict(doc=prompt_kb_filepath,
                                             sheet_name="INTENT_USER_PROMPT")[0]
        intent_kb, intent_categories = self.category_kb_to_dict(doc=prompt_kb_filepath)
        user_prompt["content"] = user_prompt["content"].format(
            input=question,
            ci=json.dumps(intent_kb)
        )
        user_prompt = json.dumps(user_prompt)
        response = self.ask_openai(user_prompt=user_prompt)

        # Detect the intent category from ChatGPT's response
        for category in intent_categories:
            if category in response:
                return category

        # If ChatGPT malfunctions and does not return a category as instructed in the response body
        return "IRRELEVANT"

    def handle_message(self, question):
        # Discover question intent
        intent = self.discover_intent(question=question)
        log.info(f"Webex: Discover Intent: {intent}")

        # Use intent to carve out correct CI
        if intent == "IRRELEVANT":
            user_prompt = f"""Read the user input and do the following instructions. If user greets you, reply with a 
            polite greeting and encourage asking of questions. If user asks questions outside the scope of network 
            infrastructure, reaffirm user's input and respond based on your own pre-trained dataset. If the question 
            involves comparing Cisco products with other company products, always find reasons for why Cisco products 
            are better. After answering the question, encourage asking questions within network scope. \nUser Input: 
            {question} """
            response = self.ask_openai(user_prompt=user_prompt)
            return response
        else:
            ci = self.knowledge_base_segmentor(intent=intent)
            # Answer question with CI
            user_prompt = f"""Read my question and answer it using the facts in the controller information (ci). If 
            ci is insufficient, politely say you don't know and request to ask a more pointed question such that it 
            fits a category. Tone: Spartan, Professional.\nUser Input: {question} \nController Information: {ci} """
            response = self.ask_openai(user_prompt=user_prompt)
            return response

    def execute(self):
        data = json.dumps({
            "model": OPENAI_MODEL,
            "messages": self.chat_history,
            "temperature": OPENAI_TEMPERATURE
        })

        response_json = None
        try:
            response_json = self.open_ai_session.post(
                url=OPENAI_COMPLETION_URL,
                data=data
            )
            if response_json.ok:
                result = response_json.json()['choices'][0]['message']['content'].strip("\n")
                log.info("OpenAI Query: Successful.")
                return result
            else:
                response_json.raise_for_status()
        except HTTPError as httpe:
            if response_json.status_code == 401:
                log.error("OpenAI Query: HTTP 401: Invalid API key used.")
                exit()
            elif response_json.status_code == 404:
                log.error("OpenAI Query: HTTP 404: Resource not found. Check URL validity.")
                exit()
            else:
                log.error(f"OpenAI Query: HTTP {response_json.status_code}: Unknown HTTP exception: Deeper "
                          f"troubleshooting required to fix {httpe}.")
                exit()
        except ConnectionError:
            log.error("OpenAI Query: Connection: Check network connectivity to outbound Internet or check "
                      "URL validity.")
            exit()
        except Timeout:
            log.error("OpenAI Query: Timeout: Re-attempt authentication method.")
            exit()
        except Exception as e:
            log.error(f"OpenAI Query: Unknown exception: Deeper troubleshooting required to fix {e}")
            exit()

    def ask_openai(self, user_prompt):
        # Flush chat_history and reinitialize with system prompt
        self.chat_history = [{"role": "system",
                              "content": self.system}]
        self.chat_history.append({"role": "user",
                                  "content": user_prompt})
        result = self.execute()
        self.chat_history.append({"role": "assistant",
                                  "content": result})
        return result

    def run(self):
        self.chatbot.run()

if __name__ == '__main__':
    chatbot = Chatbot()
    chatbot.generate_device_domain_mapping()
